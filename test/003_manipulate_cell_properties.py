import openpyxl

# import methods for manipulating alignment properties.
from openpyxl.styles import Alignment

# import utility to convert column index to letter.
import openpyxl.utils.cell  

# col index location for first parameter = 5.
COL_PARAM_FIRST = 5

# row index location for parameter descriptions.
ROW_PARAM_DESC = 4

# open workbook.
filepath = 'res\LST-200219-SLP_CONFIG-TM01-Mix.xlsx'
wb = openpyxl.load_workbook(filepath)

# select worksheet.
ws = wb['P_ANALOG_INPUT']

# set column width.
for col in range(COL_PARAM_FIRST, ws.max_column):

    # get column letter from column index.
    col_letter = openpyxl.utils.cell.get_column_letter(col)

    # set column width; openpyxl requires cols to be referenced by letter. 
    ws.column_dimensions[col_letter].width = 30

    # set text wrapping for cells.
    ws.cell(row=ROW_PARAM_DESC, column=col).alignment = Alignment(wrap_text=True)

# save and close workbook
wb.save(filepath)
wb.close()