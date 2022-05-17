import openpyxl

""" COMMENTS
- objective is to read all worksheet names from existing workbook.
- openpyxl tutorials online show the following for saving a workbook:
    
    wb = Workbook()
    wb.save(<filepath>)

  Make sure NOT include 'wb = Workbook()' as this will create a brand new workbook
  and overwrite your file!

- openpyxl takes some time to save workbooks, make sure to allow enough time! You 
  will see errors if you attempt to open the file before the save process has 
  completed.
"""

# open existing workbook
filepath = 'res\\LST-200219-SLP_CONFIG-TM01-Mix.xlsx'
wb = openpyxl.load_workbook(filepath)

# # print worksheet names in workbook
#print(wb.sheetnames)

# select worksheet by name
ws = wb['P_ANALOG_INPUT']

try:
    # write to cell in worksheet
    #ws['B8'] = 'TEST'
    ws.cell(row=8, column=2, value='TEST')

    # read from cell in worksheet
    #c = ws['B8']
    c = ws.cell(row=8, column=2)
    print(c.value)

except:
    print('Error')

# save changes to workbook
wb.save('res\\LST-200219-SLP_CONFIG-TM01-Mix.xlsx')

# close workbook
wb.close()