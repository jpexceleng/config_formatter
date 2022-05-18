"""
Extracts all param names from every worksheet in a CONFIG workbook. Writes
results to a new workbook.
"""

import openpyxl
import openpyxl.utils.cell
from openpyxl import Workbook

# tuple of worksheet target names to scan for.
WKSH_TGTS = (
    'P_ANALOG_INPUT',
    # 'P_ANALOG_OUTPUT',
    # 'P_DISCRETE_INPUT',
    # 'P_DISCRETE_OUTPUT',
    # 'P_DOSING',
    # 'P_INTERLOCK',
    # 'P_MOTOR_DISCRETE',
    # 'P_PERMISSIVE',
    # 'P_PID',
    # 'P_VALVE_DISCRETE',
    # 'P_VARIABLE_SPEED_DRIVE',
    # 'raP_Opr_Area',
    # 'raP_Opr_EMGen',
    # 'raP_Opr_EPGen',
    # 'raP_Opr_ExtddAlm',
    # 'raP_Opr_Prompt',
    # 'raP_Opr_Unit',
    # 'raP_Tec_ParRpt'
)

# start row indices for params.
ROW_PARAM_PREFIX = 7
ROW_PARAM_NAME = 8

# start col indices for params.
COL_PARAM_START = 5

# Open reference workbook for reading only.
wb_ref = openpyxl.load_workbook('res\StandardCONFIG.xlsx')

# create new workbook for writing.
wb = Workbook()

# for each worksheet, check if title matches list; if True, read param names
# from reference workbook worksheet and write to new workbook.
for ws_ref in wb_ref.worksheets:
    if ws_ref.title in WKSH_TGTS:
        # read values from cols into list
        params = []
        for col in ws_ref.iter_cols(min_row=ROW_PARAM_PREFIX, min_col = COL_PARAM_START, \
            max_col=ws_ref.max_column, max_row=ROW_PARAM_NAME, values_only=True):
            
            param = str()
            for val in col:
                # handle empty cells.
                if val is None:
                    val = ''
                # concatenate all row values in col
                param += val
            
            # only append non-blank param values
            if param != '':
                params.append(param)
        
        # create new worksheet and write params
        ws = wb.create_sheet(ws_ref.title)
        r = 1
        for param in params:
            ws.cell(row=r, column=1, value=param)
            r += 1


# clean up
wb.save('generated_workbook.xlsx')
wb.close()

wb_ref.close()