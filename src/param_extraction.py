"""
Extracts all param names from every worksheet in a CONFIG workbook. Writes
results to a new workbook.
"""
import os
import openpyxl
import openpyxl.utils.cell
from openpyxl import Workbook

# tuple of worksheet target names to scan for.
WKSH_TGTS = (
    'P_ANALOG_INPUT',
    'P_ANALOG_OUTPUT',
    'P_DISCRETE_INPUT',
    'P_DISCRETE_OUTPUT',
    'P_DOSING',
    'P_INTERLOCK',
    'P_MOTOR_DISCRETE',
    'P_PERMISSIVE',
    'P_PID',
    'P_VALVE_DISCRETE',
    'P_VARIABLE_SPEED_DRIVE',
    'raP_Opr_Area',
    'raP_Opr_EMGen',
    'raP_Opr_EPGen',
    'raP_Opr_ExtddAlm',
    'raP_Opr_Prompt',
    'raP_Opr_Unit',
    'raP_Tec_ParRpt'
)

# start row indices for params.
ROW_PARAM_PREFIX = 7
ROW_PARAM_NAME = 8

# start col indices for params.
COL_PARAM_START = 5

# workbook filepaths.
# WB_REF_FILEPATH = 'res\StandardCONFIG.xlsx'
WB_REF_FILEPATH = 'res\LST-200219-SLP_CONFIG-TM01-Mix (Rev 2)_2.xlsx'
WB_FILEPATH = 'res\compare_params.xlsx'

# Open reference workbook.
wb_ref = openpyxl.load_workbook(WB_REF_FILEPATH)

# check if workbook exists in filesystem.
if os.path.exists(WB_FILEPATH):
    # Open existing workbook for writing.
    wb = openpyxl.load_workbook(WB_FILEPATH)
else:
    # create new workbook for writing.
    wb = Workbook()

# for each worksheet, check if title matches list; if True, read param names
# from reference workbook worksheet and write to new workbook.
for ws_ref in wb_ref.worksheets:
    if ws_ref.title in WKSH_TGTS:
        # read values from cols into list.
        params = []
        for col in ws_ref.iter_cols(min_row=ROW_PARAM_PREFIX, min_col = COL_PARAM_START, \
        max_col=ws_ref.max_column, max_row=ROW_PARAM_NAME, values_only=True):
            
            param = str()
            for val in col:
                # handle empty cells.
                if val is None:
                    val = ''
                # concatenate all row values in col
                param += val
            
            # only append non-blank param values
            if param != '':
                params.append(param)
        
        # check if worksheet already exists

        # create new worksheet and write params
        if ws_ref.title in wb.sheetnames:
            ws = wb[ws_ref.title]
        else:
            ws = wb.create_sheet(ws_ref.title)

        # search left to right for next empty column.
        next_empty_col = 1
        while ws.cell(row=1, column=c).value is not None:
            next_empty_col += 1

        # write params to empty column.
        r = 1
        for param in params:
            ws.cell(row=r, column=next_empty_col, value=param)
            r += 1

# clean up
wb_ref.close()
wb.save(WB_FILEPATH)
wb.close()